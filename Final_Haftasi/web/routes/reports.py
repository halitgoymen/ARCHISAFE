import io
import json
from datetime import datetime, date, timedelta
from flask import Blueprint, render_template, request, send_file, jsonify
from flask_login import login_required
from sqlalchemy import func
from .. import db
from ..models.personnel import Personnel
from ..models.camera import Camera
from ..models.event import AttendanceLog, PPEViolation, FallEvent

reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/')
@login_required
def index():
    today = date.today()
    start_default = (today - timedelta(days=30)).isoformat()
    end_default = today.isoformat()

    start_str = request.args.get('start', start_default)
    end_str = request.args.get('end', end_default)

    try:
        start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    except ValueError:
        start_date = today - timedelta(days=30)
        end_date = today

    # Daily entry data for chart
    daily_entries = []
    cur = start_date
    while cur <= end_date:
        count = AttendanceLog.query.filter(
            func.date(AttendanceLog.entry_time) == cur
        ).count()
        daily_entries.append({'date': cur.isoformat(), 'count': count})
        cur += timedelta(days=1)

    # PPE violations in range
    ppe_violations = PPEViolation.query.filter(
        func.date(PPEViolation.timestamp) >= start_date,
        func.date(PPEViolation.timestamp) <= end_date,
    ).order_by(PPEViolation.timestamp.desc()).limit(100).all()

    # Fall events in range
    fall_events = FallEvent.query.filter(
        func.date(FallEvent.timestamp) >= start_date,
        func.date(FallEvent.timestamp) <= end_date,
    ).order_by(FallEvent.timestamp.desc()).all()

    # Top violators
    top_violators = db.session.query(
        Personnel,
        func.count(PPEViolation.id).label('violation_count')
    ).join(PPEViolation, Personnel.id == PPEViolation.personnel_id).filter(
        func.date(PPEViolation.timestamp) >= start_date,
        func.date(PPEViolation.timestamp) <= end_date,
    ).group_by(Personnel.id).order_by(
        func.count(PPEViolation.id).desc()
    ).limit(10).all()

    summary = {
        'total_entries': AttendanceLog.query.filter(
            func.date(AttendanceLog.entry_time) >= start_date,
            func.date(AttendanceLog.entry_time) <= end_date,
        ).count(),
        'total_ppe_violations': len(ppe_violations),
        'total_fall_events': len(fall_events),
        'active_personnel': Personnel.query.filter_by(is_active=True).count(),
    }

    return render_template(
        'reports/index.html',
        start_date=start_date,
        end_date=end_date,
        daily_entries=daily_entries,
        ppe_violations=ppe_violations,
        fall_events=fall_events,
        top_violators=top_violators,
        summary=summary,
    )


@reports_bp.route('/export/excel')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    start_str = request.args.get('start', '')
    end_str = request.args.get('end', '')
    report_type = request.args.get('type', 'attendance')

    try:
        start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    except ValueError:
        end_date = date.today()
        start_date = end_date - timedelta(days=30)

    wb = openpyxl.Workbook()
    ws = wb.active

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')

    if report_type == 'attendance':
        ws.title = 'Giriş Çıkış Raporu'
        headers = ['Tarih', 'Personel ID', 'Ad Soyad', 'Görev', 'Giriş Saati',
                   'Çıkış Saati', 'Süre (saat)', 'Yöntem']
        logs = AttendanceLog.query.filter(
            func.date(AttendanceLog.entry_time) >= start_date,
            func.date(AttendanceLog.entry_time) <= end_date,
        ).order_by(AttendanceLog.entry_time.desc()).all()

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for row, log in enumerate(logs, 2):
            p = log.personnel
            ws.cell(row=row, column=1, value=log.entry_time.date().isoformat())
            ws.cell(row=row, column=2, value=p.personnel_id if p else '')
            ws.cell(row=row, column=3, value=p.full_name if p else '')
            ws.cell(row=row, column=4, value=p.job_title if p else '')
            ws.cell(row=row, column=5,
                    value=log.entry_time.strftime('%H:%M:%S'))
            ws.cell(row=row, column=6,
                    value=log.exit_time.strftime('%H:%M:%S') if log.exit_time else 'Açık')
            ws.cell(row=row, column=7, value=log.duration_hours)
            ws.cell(row=row, column=8, value=log.method)

    elif report_type == 'ppe':
        ws.title = 'KKD İhlal Raporu'
        headers = ['Tarih/Saat', 'Personel ID', 'Ad Soyad', 'Görev',
                   'Eksik Ekipmanlar', 'Tespit Edilen', 'Güven Skoru', 'Durum']
        violations = PPEViolation.query.filter(
            func.date(PPEViolation.timestamp) >= start_date,
            func.date(PPEViolation.timestamp) <= end_date,
        ).order_by(PPEViolation.timestamp.desc()).all()

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for row, v in enumerate(violations, 2):
            p = v.personnel
            ws.cell(row=row, column=1, value=v.timestamp.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=2, value=p.personnel_id if p else '')
            ws.cell(row=row, column=3, value=p.full_name if p else 'Bilinmiyor')
            ws.cell(row=row, column=4, value=p.job_title if p else '')
            ws.cell(row=row, column=5, value=', '.join(v.missing_ppe))
            ws.cell(row=row, column=6, value=', '.join(v.detected_ppe))
            ws.cell(row=row, column=7, value=f'{(v.confidence or 0):.0%}')
            ws.cell(row=row, column=8, value='Çözüldü' if v.is_resolved else 'Açık')

    elif report_type == 'fall':
        ws.title = 'Düşme Olayları Raporu'
        headers = ['Tarih/Saat', 'Kamera', 'Konum', 'Güven Skoru',
                   'Personel', 'Video', 'Durum']
        events = FallEvent.query.filter(
            func.date(FallEvent.timestamp) >= start_date,
            func.date(FallEvent.timestamp) <= end_date,
        ).order_by(FallEvent.timestamp.desc()).all()

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for row, e in enumerate(events, 2):
            ws.cell(row=row, column=1, value=e.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=2, value=e.camera.name if e.camera else '')
            ws.cell(row=row, column=3, value=e.location_desc)
            ws.cell(row=row, column=4, value=f'{(e.confidence or 0):.0%}')
            ws.cell(row=row, column=5, value=e.personnel.full_name if e.personnel else '')
            ws.cell(row=row, column=6, value=e.video_path or '')
            ws.cell(row=row, column=7, value='Çözüldü' if e.is_resolved else 'Açık')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'archisafe_{report_type}_{start_date}_{end_date}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument'
                                     '.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)
